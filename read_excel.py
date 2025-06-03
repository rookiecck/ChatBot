import openpyxl

def read_all_sheets(file_path):
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # 获取所有工作表的名称
        sheet_names = workbook.sheetnames

        # 存储每个工作表的数据的字典
        all_data = {}

        # 遍历每个工作表并读取数据
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            all_data[sheet_name] = data

        return all_data

    except Exception as e:
        return str(e)

# Excel 文件路径
excel_file_path = 'D:/企业微信下载文件/WXWorkLocal/File/2023-08/is_value_in_range.xlsx'

# 读取所有工作表的内容
all_sheets_data = read_all_sheets(excel_file_path)

if isinstance(all_sheets_data, dict):
    for sheet_name, data in all_sheets_data.items():
        print(f"Sheet: {sheet_name}")
        for row in data:
            print(row)
        print("-" * 30)
else:
    print("Error:", all_sheets_data)